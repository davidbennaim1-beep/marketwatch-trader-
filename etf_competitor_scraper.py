#!/usr/bin/env python3
"""
Defiance ETF Competitor Analysis
=================================
Finds top 5 competitors for each Defiance ETF, compares AUM, expense ratio,
and launch date, then exports a formatted Excel report.

Data sources:
  - stockanalysis.com  : ETF details (AUM, expense ratio, inception date, category)
  - etfdb.com          : Competitor discovery (category-based "Other ETFs" section)

Usage:
  python3 etf_competitor_scraper.py              # full run (all ~80 ETFs)
  python3 etf_competitor_scraper.py --test       # test run (first 5 ETFs only)
  python3 etf_competitor_scraper.py --clear-cache  # wipe cached data and re-fetch
"""

import re
import json
import time
import os
import sys
import argparse
from datetime import datetime

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

INPUT_PATH  = '/Users/david/Downloads/Defiance Etfs list.xlsx'
OUTPUT_PATH = '/Users/david/Downloads/Defiance_Competitor_Analysis.xlsx'
CACHE_FILE  = '/tmp/etf_scraper_cache.json'

REQUEST_DELAY = 1.5   # seconds between HTTP requests
MAX_CANDIDATES = 30   # max competitor candidates to look up per ETF

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/122.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}

# Category keywords that indicate a money market / fixed income / clearly unrelated fund
EXCLUDED_CATEGORIES = {
    'money market', 'government bond', 'fixed income', 'bond', 'treasury',
    'municipal', 'corporate bond', 'inflation-protected', 'ultrashort',
    'short-term bond', 'intermediate bond', 'long-term bond', 'currency',
    'preferred stock', 'real estate', 'commodities', 'commodity',
    'volatility', 'multi-asset', 'allocation', 'balanced',
}

# Words that appear after "2X Long" in ETF names but are NOT stock tickers
NON_TICKER_WORDS = {
    'uranium', 'gold', 'oil', 'silver', 'bitcoin', 'ethereum',
    'pure', 'quantum', 'gen', 'next', 'long', 'short', 'daily',
    'target', 'tech', 'china', 'india', 'europe', 'japan',
    'sector', 'index', 'fund', 'inc', 'llc', 'corp', 'crypto',
    'blockchain', 'hydrogen', 'aerospace', 'defense', 'retail',
    'hotel', 'airline', 'cruise', 'nuclear', 'clean', 'energy',
    'ai', 'mag', 'trillion', 'connective', 'r2000', 'nasdaq',
    'income', 'distribution', 'options', 'volatility',
}


# ─────────────────────────────────────────────────────────────────────────────
# CACHE
# ─────────────────────────────────────────────────────────────────────────────

def load_cache() -> dict:
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_cache(cache: dict):
    try:
        with open(CACHE_FILE, 'w') as f:
            json.dump(cache, f, indent=2)
    except Exception as e:
        print(f'  [cache] save error: {e}')


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def aum_to_millions(s) -> float:
    """Convert '$3.48B', '$665.67M', '$998.54K', '$173.0 M' → float in millions."""
    if not s:
        return 0.0
    s = re.sub(r'[,$\s]', '', str(s)).upper()
    m = re.match(r'([\d.]+)([BMK]?)', s)
    if not m:
        return 0.0
    val, unit = float(m.group(1)), m.group(2)
    return val * {'B': 1000, 'M': 1, 'K': 0.001}.get(unit, 1e-6)


def fmt_aum(m: float) -> str:
    if m >= 1000:  return f'${m / 1000:.2f}B'
    if m >= 1:     return f'${m:.2f}M'
    if m > 0:      return f'${m * 1000:.2f}K'
    return 'N/A'


def fmt_pct(ratio) -> str:
    if ratio is None:
        return 'N/A'
    try:
        return f'{float(ratio) * 100:.2f}%'
    except Exception:
        return 'N/A'


def get_html(url: str, cache_key: str, cache: dict) -> str | None:
    """GET a URL, with caching. Returns HTML string or None."""
    if cache_key in cache:
        return cache[cache_key]
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        time.sleep(REQUEST_DELAY)
        if r.status_code != 200:
            return None
        cache[cache_key] = r.text
        save_cache(cache)
        return r.text
    except Exception as e:
        print(f'    [HTTP] {url}: {e}')
        time.sleep(REQUEST_DELAY)
        return None


# ─────────────────────────────────────────────────────────────────────────────
# LOAD DEFIANCE ETF LIST FROM EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def load_defiance_etfs() -> list[dict]:
    wb = openpyxl.load_workbook(INPUT_PATH)
    ws = wb.active
    etfs, started = [], False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'Ticker':
            started = True
            continue
        if started and row[0] and isinstance(row[0], str):
            etfs.append({
                'ticker':        row[0].strip(),
                'name':          str(row[1] or '').strip(),
                'aum_millions':  aum_to_millions(row[2]),
                'expense_ratio': float(row[3]) if row[3] else None,
            })
    return etfs


# ─────────────────────────────────────────────────────────────────────────────
# stockanalysis.com — ETF details
# ─────────────────────────────────────────────────────────────────────────────

def sa_get_etf_details(ticker: str, cache: dict) -> dict | None:
    """
    Scrape stockanalysis.com for an ETF's details.
    Returns dict: {ticker, name, aum_millions, expense_ratio, inception_date, category}
    """
    html = get_html(
        f'https://stockanalysis.com/etf/{ticker.lower()}/',
        f'sa_page_{ticker.upper()}',
        cache
    )
    if not html:
        return None

    soup = BeautifulSoup(html, 'html.parser')
    text = soup.get_text(separator=' ')
    data = {'ticker': ticker.upper()}

    # AUM and Expense Ratio — first stats table
    for table in soup.find_all('table'):
        for tr in table.find_all('tr'):
            cells = [td.get_text(strip=True) for td in tr.find_all(['td', 'th'])]
            if len(cells) >= 2:
                label = cells[0].strip().lower()
                val   = cells[1].strip()
                if label == 'assets' and 'aum_millions' not in data:
                    data['aum_str']     = val
                    data['aum_millions'] = aum_to_millions(val)
                elif label == 'expense ratio' and 'expense_ratio' not in data:
                    data['expense_ratio_str'] = val
                    try:
                        data['expense_ratio'] = float(val.replace('%', '')) / 100
                    except Exception:
                        pass

    # Inception / Launch date
    m = re.search(r'Inception\s*Date\s+([A-Z][a-z]{2}\s+\d{1,2},?\s+\d{4})', text)
    if m:
        data['inception_date'] = m.group(1).replace(',', '')
    else:
        m2 = re.search(r'Inception\s*Date\s+(\d{2}/\d{2}/\d{4})', text)
        data['inception_date'] = m2.group(1) if m2 else 'N/A'

    # Category
    m = re.search(
        r'Category\s+([A-Za-z &/\-]+?)(?:Stock Exchange|Ticker Symbol|ETF Provider|Index)',
        text
    )
    data['category'] = m.group(1).strip() if m else ''

    # Name from h1 (strip ticker in parentheses)
    h1 = soup.find('h1')
    if h1:
        name = h1.get_text(strip=True)
        name = re.sub(r'\s*\([A-Z]{1,5}\)\s*$', '', name)
        data['name'] = name.strip()
    else:
        data['name'] = ticker.upper()

    return data


# ─────────────────────────────────────────────────────────────────────────────
# etfdb.com — Competitor discovery
# ─────────────────────────────────────────────────────────────────────────────

def etfdb_get_competitors(ticker: str, cache: dict) -> list[str]:
    """
    Scrape etfdb.com ETF page for competitor/alternative ETF tickers.
    Focuses on the 'Other ETFs in the ETF Database Category' section
    and 'Alternative ETFs' sections.
    """
    html = get_html(
        f'https://etfdb.com/etf/{ticker.upper()}/',
        f'etfdb_page_{ticker.upper()}',
        cache
    )
    if not html:
        return []

    soup = BeautifulSoup(html, 'html.parser')
    found: list[str] = []

    # Priority 1: "Other ETFs in the ETF Database Category" — most relevant section
    for heading in soup.find_all(['h2', 'h3', 'h4']):
        txt = heading.get_text(strip=True)
        if 'Other ETFs in the ETF Database' in txt or 'Alternative ETFs' in txt:
            # Walk siblings until next heading of same or higher level
            sibling = heading.find_next_sibling()
            while sibling and sibling.name not in ('h2', 'h3', 'h4'):
                for a in sibling.find_all('a', href=re.compile(r'/etf/[A-Z]{1,5}/')):
                    t = a['href'].strip('/').split('/')[-1].upper()
                    if t and re.match(r'^[A-Z]{1,5}$', t) and t != ticker.upper():
                        found.append(t)
                # Also check text directly for ticker symbols
                for span_or_td in sibling.find_all(['td', 'span', 'a']):
                    text = span_or_td.get_text(strip=True).upper()
                    if re.match(r'^[A-Z]{1,5}$', text) and text != ticker.upper():
                        found.append(text)
                sibling = sibling.find_next_sibling()

    # Priority 2: All ETF links on the page (broader catch)
    for a in soup.find_all('a', href=re.compile(r'/etf/[A-Z]{1,5}/$')):
        t = a['href'].strip('/').split('/')[-1].upper()
        if t and t != ticker.upper():
            found.append(t)

    return list(dict.fromkeys(found))  # deduplicate, preserve order


# ─────────────────────────────────────────────────────────────────────────────
# stockanalysis.com — additional similar ETF discovery
# ─────────────────────────────────────────────────────────────────────────────

def sa_get_similar_tickers(ticker: str, cache: dict) -> list[str]:
    """Get linked ETF tickers from the stockanalysis.com ETF page."""
    html = cache.get(f'sa_page_{ticker.upper()}')  # reuse cached page if available
    if not html:
        html = get_html(
            f'https://stockanalysis.com/etf/{ticker.lower()}/',
            f'sa_page_{ticker.upper()}',
            cache
        )
    if not html:
        return []

    soup = BeautifulSoup(html, 'html.parser')
    found = []
    for a in soup.find_all('a', href=re.compile(r'^/etf/[a-z]+/?$')):
        t = a['href'].strip('/').split('/')[-1].upper()
        if t and re.match(r'^[A-Z]{1,5}$', t) and t != ticker.upper():
            found.append(t)
    return list(dict.fromkeys(found))


# ─────────────────────────────────────────────────────────────────────────────
# UNDERLYING TICKER EXTRACTION  (for single-stock leveraged ETFs)
# ─────────────────────────────────────────────────────────────────────────────

_LEV_RE = re.compile(
    r'(?:Daily\s+Target\s+)?(?:2[Xx]|3[Xx])\s+(?:Daily\s+)?(?:Long|Short)\s+([A-Z]{2,5})\s+ETF',
    re.IGNORECASE
)
_LEV_INCOME_RE = re.compile(
    r'Leveraged\s+(?:Long\s+)?(?:\+\s+)?(?:Income\s+)?([A-Z]{2,5})\s+ETF',
    re.IGNORECASE
)


def extract_underlying(name: str) -> str | None:
    for pattern in (_LEV_RE, _LEV_INCOME_RE):
        m = pattern.search(name)
        if m:
            candidate = m.group(1).upper()
            if candidate.lower() not in NON_TICKER_WORDS and len(candidate) >= 2:
                return candidate
    return None


# ─────────────────────────────────────────────────────────────────────────────
# MAIN COMPETITOR FINDER
# ─────────────────────────────────────────────────────────────────────────────

def find_competitors(
    etf: dict,
    defiance_details: dict,
    cache: dict,
    all_defiance_tickers: set,
    num: int = 5,
) -> list[dict]:
    """
    Find top `num` competitor ETFs for a Defiance ETF.

    Discovery order:
      1. etfdb.com similar/alternative ETF links (best structured source)
      2. stockanalysis.com ETF page links
    Then fetch stockanalysis.com details for each candidate,
    filter out Defiance ETFs, sort by AUM, return top `num`.
    For single-stock leveraged ETFs, prefer competitors that share
    the same underlying stock in their name.
    """
    ticker     = etf['ticker']
    name       = etf['name']
    underlying = extract_underlying(name)

    # --- Collect candidates ---
    candidates: list[str] = []

    etfdb_tickers = etfdb_get_competitors(ticker, cache)
    candidates += etfdb_tickers

    sa_tickers = sa_get_similar_tickers(ticker, cache)
    candidates += sa_tickers

    # Deduplicate, remove self and all Defiance ETFs
    seen: set = set()
    clean: list[str] = []
    for t in candidates:
        t = t.upper()
        if t not in seen and t != ticker.upper() and t not in all_defiance_tickers:
            seen.add(t)
            clean.append(t)

    print(f'    {len(clean)} unique candidates')

    # --- Fetch details for each candidate ---
    competitors: list[dict] = []
    for t in clean[:MAX_CANDIDATES]:
        details = sa_get_etf_details(t, cache)
        if details and details.get('aum_millions', 0) > 0:
            competitors.append(details)

    # Filter out ETFs in clearly unrelated categories (e.g. money market vs equity)
    def is_compatible(comp: dict) -> bool:
        cat = comp.get('category', '').lower()
        return not any(excl in cat for excl in EXCLUDED_CATEGORIES)

    competitors = [c for c in competitors if is_compatible(c)]

    # Sort by AUM descending
    competitors.sort(key=lambda x: x.get('aum_millions', 0), reverse=True)

    # For single-stock leveraged ETFs: ONLY keep competitors that share the
    # same underlying stock. These are often one-of-a-kind — if no true direct
    # competitor exists, return empty rather than filling with unrelated ETFs.
    if underlying:
        competitors = [
            c for c in competitors
            if underlying.upper() in c.get('name', '').upper()
            or underlying.upper() in c.get('ticker', '').upper()
        ]

    return competitors[:num]


# ─────────────────────────────────────────────────────────────────────────────
# BUILD RESULT ROW
# ─────────────────────────────────────────────────────────────────────────────

def build_result(defiance_etf: dict, defiance_details: dict, competitors: list[dict]) -> dict:
    all_etfs = [defiance_details] + competitors

    # AUM rank (1 = largest)
    sorted_by_aum = sorted(all_etfs, key=lambda x: x.get('aum_millions', 0), reverse=True)
    rank = next(
        (i + 1 for i, e in enumerate(sorted_by_aum)
         if e.get('ticker', '').upper() == defiance_etf['ticker'].upper()),
        'N/A'
    )

    # Lowest expense ratio in the peer group + which ticker holds it
    etfs_with_exp = [(e['expense_ratio'], e.get('ticker', ''))
                     for e in all_etfs if e.get('expense_ratio') is not None]
    if etfs_with_exp:
        lowest_exp, lowest_exp_ticker = min(etfs_with_exp, key=lambda x: x[0])
    else:
        lowest_exp, lowest_exp_ticker = None, ''

    # First mover: did Defiance launch before ALL competitors?
    def parse_date(s):
        if not s or s == 'N/A':
            return None
        for fmt in ('%b %d %Y', '%b %d, %Y', '%m/%d/%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(s.strip(), fmt)
            except ValueError:
                pass
        return None

    def_date = parse_date(defiance_details.get('inception_date'))
    comp_dates = [parse_date(c.get('inception_date')) for c in competitors]
    comp_dates = [d for d in comp_dates if d is not None]

    if def_date and comp_dates:
        first_mover = def_date < min(comp_dates)
    else:
        first_mover = None  # can't determine

    return {
        'defiance':          defiance_details,
        'competitors':       competitors,
        'rank':              rank,
        'lowest_exp':        lowest_exp,
        'lowest_exp_ticker': lowest_exp_ticker,
        'first_mover':       first_mover,
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

C_BLUE_DARK  = 'FF1E3A5F'
C_BLUE_LIGHT = 'FFD6E4F7'
C_WHITE      = 'FFFFFFFF'
C_STRIPE     = 'FFF2F7FD'
C_HEADER_BG  = 'FF2B2B2B'
C_GOLD       = 'FFFFD700'
C_SEPARATOR  = 'FFDDDDDD'
C_TEXT       = 'FF1A1A1A'
C_MUTED      = 'FF888888'


def _border(color='FFBBBBBB'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value='', bg=C_WHITE, fg=C_TEXT,
          bold=False, align='left', wrap=False, bc='FFBBBBBB'):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill('solid', fgColor=bg)
    c.font      = Font(name='Calibri', size=10, bold=bold, color=fg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border    = _border(bc)
    return c


COLUMNS = [
    ('Type',                              13),
    ('Ticker',                             9),
    ('Fund Name',                         46),
    ('AUM',                               13),
    ('Expense\nRatio',                    12),
    ('Launch Date',                       13),
    ('AUM Rank\n(among peers)',           14),
    ('Lowest Expense\nRatio in Group',    22),
    ('First\nMover?',                     10),
]

C_GREEN_BG = 'FFD6F5D6'
C_RED_BG   = 'FFFFD6D6'
C_GREEN_FG = 'FF1A6B1A'
C_RED_FG   = 'FF8B0000'


def write_block(ws, start_row: int, result: dict) -> int:
    d           = result['defiance']
    cps         = result['competitors']
    rnk         = result['rank']
    lwe         = result['lowest_exp']
    lwe_ticker  = result.get('lowest_exp_ticker', '')
    first_mover = result.get('first_mover')
    row         = start_row

    # Lowest expense ratio cell value: "0.08% (XLK)"
    lwe_str = f'{fmt_pct(lwe)} ({lwe_ticker})' if lwe_ticker else fmt_pct(lwe)

    # First mover display
    if first_mover is True:
        fm_val, fm_bg, fm_fg = 'Yes', C_GREEN_BG, C_GREEN_FG
    elif first_mover is False:
        fm_val, fm_bg, fm_fg = 'No', C_RED_BG, C_RED_FG
    else:
        fm_val, fm_bg, fm_fg = 'N/A', C_BLUE_DARK, C_WHITE

    # ── Defiance row (dark blue) ─────────────────────────────────────────────
    rank_str = f'#{rnk} of {len(cps) + 1}' if isinstance(rnk, int) else 'N/A'
    row_vals = [
        ('Defiance ETF',                    'center', C_BLUE_DARK, C_WHITE),
        (d.get('ticker', ''),               'center', C_BLUE_DARK, C_WHITE),
        (d.get('name', ''),                 'left',   C_BLUE_DARK, C_WHITE),
        (fmt_aum(d.get('aum_millions', 0)), 'right',  C_BLUE_DARK, C_WHITE),
        (fmt_pct(d.get('expense_ratio')),   'center', C_BLUE_DARK, C_WHITE),
        (d.get('inception_date', 'N/A'),    'center', C_BLUE_DARK, C_WHITE),
        (rank_str,                          'center', C_BLUE_DARK, C_GOLD),
        (lwe_str,                           'center', C_BLUE_DARK, C_GOLD),
        (fm_val,                            'center', fm_bg,        fm_fg),
    ]
    for col, (val, align, bg, fg) in enumerate(row_vals, 1):
        _cell(ws, row, col, val, bg=bg, fg=fg, bold=True, align=align, bc='FF1E3A5F')
    ws.row_dimensions[row].height = 19
    row += 1

    # ── Competitor rows (only write rows where a competitor was found) ────────
    for i, cp in enumerate(cps):
        bg = C_BLUE_LIGHT if i % 2 == 0 else C_STRIPE
        vals = [
            (f'Competitor {i + 1}',              'center'),
            (cp.get('ticker', ''),               'center'),
            (cp.get('name', ''),                 'left'),
            (fmt_aum(cp.get('aum_millions', 0)), 'right'),
            (fmt_pct(cp.get('expense_ratio')),   'center'),
            (cp.get('inception_date', 'N/A'),    'center'),
            ('', 'center'),
            ('', 'center'),
            ('', 'center'),
        ]
        for col, (val, align) in enumerate(vals, 1):
            _cell(ws, row, col, val, bg=bg, fg=C_TEXT, align=align)
        ws.row_dimensions[row].height = 17
        row += 1

    # ── Spacer ────────────────────────────────────────────────────────────────
    for col in range(1, len(COLUMNS) + 1):
        ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=C_SEPARATOR)
    ws.row_dimensions[row].height = 5
    row += 1

    return row


def setup_sheet(ws, title_text: str):
    """Apply title row, column headers, and formatting to a worksheet."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    # Title row
    ws.merge_cells(f'A1:{get_column_letter(len(COLUMNS))}1')
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.fill      = PatternFill('solid', fgColor=C_BLUE_DARK)
    tc.font      = Font(name='Calibri', size=13, bold=True, color=C_WHITE)
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Column headers
    for col, (label, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.fill      = PatternFill('solid', fgColor=C_HEADER_BG)
        c.font      = Font(name='Calibri', size=10, bold=True, color=C_WHITE)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        s = Side(style='medium', color='FF000000')
        c.border    = Border(left=s, right=s, top=s, bottom=s)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 30


def add_footer(ws, current_row: int, note_text: str):
    ws.merge_cells(f'A{current_row}:{get_column_letter(len(COLUMNS))}{current_row}')
    note = ws.cell(row=current_row, column=1, value=note_text)
    note.fill      = PatternFill('solid', fgColor='FFFFFBE6')
    note.font      = Font(name='Calibri', size=9, italic=True, color='FF555555')
    note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[current_row].height = 28


def export_to_excel(thematic_results: list[dict], single_stock_results: list[dict]):
    wb = openpyxl.Workbook()
    generated = datetime.now().strftime("%B %d, %Y")

    # ── Sheet 1: Thematic ETFs ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Thematic ETFs'
    setup_sheet(ws1, f'Defiance Thematic ETF Competitor Analysis  ·  {generated}')
    row = 3
    for result in thematic_results:
        row = write_block(ws1, row, result)
    add_footer(ws1, row,
        'Thematic & category ETFs — competitors sourced from the same ETF category. '
        'AUM Rank = Defiance rank by AUM among peers (1 = largest). '
        'Lowest Expense Ratio = lowest fee in the peer group. '
        'Data sourced from stockanalysis.com and etfdb.com.'
    )

    # ── Sheet 2: Single-Stock Leveraged ETFs ──────────────────────────────────
    ws2 = wb.create_sheet(title='Single-Stock Leveraged ETFs')
    setup_sheet(ws2, f'Defiance Single-Stock Leveraged ETF Competitor Analysis  ·  {generated}')
    row = 3
    for result in single_stock_results:
        row = write_block(ws2, row, result)
    add_footer(ws2, row,
        'Single-stock leveraged ETFs — competitors shown only where a true direct competitor '
        '(same underlying stock, different issuer) was found. Blank rows = no direct competitor exists. '
        'Data sourced from stockanalysis.com and etfdb.com.'
    )

    wb.save(OUTPUT_PATH)
    print(f'\n✓  Saved → {OUTPUT_PATH}')
    print(f'   Sheet 1 "Thematic ETFs":              {len(thematic_results)} ETFs')
    print(f'   Sheet 2 "Single-Stock Leveraged ETFs": {len(single_stock_results)} ETFs')


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Defiance ETF Competitor Analysis')
    parser.add_argument('--test',        action='store_true', help='Run on first 5 ETFs only')
    parser.add_argument('--clear-cache', action='store_true', help='Delete cached data')
    args = parser.parse_args()

    if args.clear_cache and os.path.exists(CACHE_FILE):
        os.remove(CACHE_FILE)
        print('Cache cleared.\n')

    print('Loading Defiance ETF list...')
    all_etfs = load_defiance_etfs()
    # Always filter out ALL Defiance ETFs from competitors, even in test mode
    all_defiance_tickers = {e['ticker'].upper() for e in all_etfs}

    if args.test:
        defiance_etfs = all_etfs[:5]
        print(f'TEST MODE: {len(defiance_etfs)} ETFs\n')
    else:
        defiance_etfs = all_etfs
        print(f'{len(defiance_etfs)} ETFs loaded\n')
    cache   = load_cache()
    results = []

    for i, etf in enumerate(defiance_etfs):
        ticker = etf['ticker']
        print(f'[{i + 1:2d}/{len(defiance_etfs)}]  {ticker:<8}  {etf["name"][:55]}')

        # Get Defiance ETF details
        details = sa_get_etf_details(ticker, cache)
        if details:
            # Prefer Excel data for AUM/expense if stockanalysis returns nothing
            if details.get('aum_millions', 0) < 0.01:
                details['aum_millions'] = etf['aum_millions']
            if details.get('expense_ratio') is None:
                details['expense_ratio'] = etf['expense_ratio']
        else:
            details = {
                'ticker':         ticker,
                'name':           etf['name'],
                'aum_millions':   etf['aum_millions'],
                'expense_ratio':  etf['expense_ratio'],
                'inception_date': 'N/A',
                'category':       '',
            }

        # Find top 5 competitors
        competitors = find_competitors(
            etf, details, cache, all_defiance_tickers
        )
        result = build_result(etf, details, competitors)
        results.append(result)

        rank_str = f'#{result["rank"]}' if isinstance(result['rank'], int) else 'N/A'
        print(f'           → {len(competitors)} competitors | AUM rank {rank_str}')

    # Split results into thematic vs single-stock
    thematic_results     = []
    single_stock_results = []
    for etf, result in zip(defiance_etfs, results):
        if extract_underlying(etf['name']):
            single_stock_results.append(result)
        else:
            thematic_results.append(result)

    print(f'\nExporting to Excel ({len(thematic_results)} thematic, {len(single_stock_results)} single-stock)...')
    export_to_excel(thematic_results, single_stock_results)


if __name__ == '__main__':
    main()
